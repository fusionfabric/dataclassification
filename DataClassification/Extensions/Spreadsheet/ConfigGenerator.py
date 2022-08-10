import sys, os, logging
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

import openpyxl, json, xlsxwriter,csv, warnings
from DataClassification.Extensions.Swaggers.SwaggerClassifier import SwaggerClassifier
warnings.simplefilter("ignore")
class JSONGenerator:
    """
            Used for grabbing rules and definitions from the dictionary template file and creating a rules.json file
            which can then be used throughout the classification process as reference for both classification rules
            and swagger definitions
            param product_directory supplies swagger folder path
            param datafile supplies path to master dictionary
            param fileTypes provides list of supported file types
            param filename provides path to template file for rules and definitions
            """

    def __init__(self,product_directory,datafile,fileTypes,filename = os.path.join(os.path.dirname(__file__),"..","..","Config/Dictionary.xlsx")):
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(filename=product_directory + '/data-classification.log', format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
        try:
            self.wb = openpyxl.load_workbook(filename)
            logging.info("Data Dictionary.xlsx file has been loaded")
            self.detailslist=self.__read_csv_file(datafile)
            logging.info("Dictionary Details.csv file has been loaded")
            self.filetypes=fileTypes
            self.definitions_list = self.get_definitions()
            self.tag_list= {}
        except Exception as e:
            logging.error(e)
            sys.exit(e)
     

    def generate_json_files(self, path = os.path.join(os.path.dirname(__file__),"..","..","Config/rules.json")):
        logging.info("Fetching definitions and rules for rules.json file")
        definitions_obj = json.dumps({
            'definitions': self.definitions_list,
            'rules' : self.get_rules()
            })
        try:
            with open(path, 'w') as outfile: 
                outfile.write(definitions_obj) 
            logging.info("rules.json file has been created")
        except Exception as e:
            logging.error(e)
            sys.exit(e) 

    def __generate_json_file(self, key, list_of_values, filename):
        definitions_obj = json.dumps({key: list_of_values})
        with open(filename, "w") as outfile: 
            outfile.write(definitions_obj) 

    def get_definitions(self):
        definitions_list = []
        try:
            field_types_sheet = self.wb["Field Types"]
            for i in range(2, field_types_sheet.max_row+1):
                if field_types_sheet.cell(row = i, column = 1).value == 'None' or field_types_sheet.cell(row = i, column = 2).value == 'None':
                    raise ValueError('Error in row #{}: Field Types and Descriptions cannot be empty'.format(i+1))
                definition = {
                    'friendly_name' : field_types_sheet.cell(row = i, column = 1).value,
                    'description' : field_types_sheet.cell(row = i, column = 2).value,
                    'examples' : field_types_sheet.cell(row = i, column = 3).value,
                    'sensitive' : field_types_sheet.cell(row = i, column = 4).value,
                    'guidelines' : field_types_sheet.cell(row = i, column = 5).value,
                    'technical_names' : self.__get_technical_names_by_friendly_name(field_types_sheet.cell(row = i, column = 1).value),
                    'Exception_API_and_Values': self.__get_exception_names_by_friendly_name(field_types_sheet.cell(row=i, column=1).value)
                }
                definitions_list.append(definition)
        except Exception as e:
            logging.error(e)
            sys.exit(e)
        return definitions_list

    def get_rules(self):
        rules_list = []
      
        try:
            rules_sheet =  self.wb['Rules']
            column_indexes = self.__get_rules_column_indexes()
            for i in range(2, rules_sheet.max_row+1):
                rule = {'any': [], 'all': [], 'tags': [], 'scope':[]}  
                n=0          
                for c in column_indexes['any']:
                    if rules_sheet.cell(row = i, column = c).value != None:
                        rule['any'].append(rules_sheet.cell(row = i, column = c).value)
                for c in column_indexes['all']:
                    if rules_sheet.cell(row = i, column = c).value != None:
                        rule['all'].append(rules_sheet.cell(row = i, column = c).value)
                for c in column_indexes['tags']:
                    n+=1
                    if rules_sheet.cell(row = i, column = c).value != None:
                        rule['tags'].append(rules_sheet.cell(row = i, column = c).value)
                        tagn='tag{}'.format(n)
                        rule['scope'].append(tagn)
                        if tagn not in self.tag_list.keys():
                            self.tag_list[tagn]=[]
                        self.tag_list[tagn].append(rules_sheet.cell(row = i, column = c).value)
            
                if (len(rule['any']) > 0 or len(rule['all']) > 0) and len(rule['tags']) > 0:
                    rules_list.append(rule)
                elif len(rule['any']) > 0 or len(rule['all']) > 0:  # No tags assigned to a rule
                    raise ValueError('Error in row #{}: Tags must be assigned to each rule'.format(i+1))
                if 'swagger' in self.filetypes:
                        rules_list=[x for x in rules_list if 'tag1' in x['scope']]
        except Exception as e:
            logging.error(e)
            sys.exit(e)
        return rules_list

    def __get_rules_column_indexes(self):
        index_dict = {'any': [], 'all': [], 'tags': [], 'scope':[]}
        rules_sheet =  self.wb['Rules']
        for i in range(1, rules_sheet.max_column+1):
            column_name = rules_sheet.cell(row = 1, column = i).value
            if 'Any' in column_name:
                index_dict['any'].append(i)
            elif 'All' in column_name:
                index_dict['all'].append(i)
            elif 'Tag' in column_name:
                if 'swagger' in self.filetypes and column_name != 'Tag2':
                    index_dict['tags'].append(i)
                elif 'swagger' not in self.filetypes:
                    index_dict['tags'].append(i)
        return index_dict

    def __get_technical_names_by_friendly_name(self, friendly_name):
        tech_names_list = []
        for item in self.detailslist[1:]:
            if friendly_name==item[0] and item[4]=='':
                tech_names_list.append(item[1])
        return list(set(tech_names_list))

    def __get_exception_names_by_friendly_name(self, friendly_name):
        exception_names_list = []
        for item in self.detailslist[1:]:
            if friendly_name==item[0]:
                if item[4]:
                    for exception_names in item[4].split(";"):
                        exception_names_list.append(exception_names+":"+item[1])
        return list(set(exception_names_list))
        
    def __read_csv_file(self,path):
        with open(path,"r",encoding='utf-8') as csvFile:
            data=csv.reader(csvFile)
            datalist=list(data)
            return datalist


# This class depends on a succesful execution of the Swagger Classifier, as it needs all classified results to generate a csv file
class SpreadsheetGenerator:
    """
                Used for creating the generated dictionary file which gives us a holistic view of the parsed out swagger
                definitions, their descriptions and classifications
                param root_folder supplies swagger folder path
                param swagger_classifier supplies artifacts generated by SwaggerClassifier class
                param json_generator supplies artifacts generated by JSONGenerator class
                param exported_file_name supplies path for creating the Generated Dictionary
                param app name provides the name of the application/product being classified
                param filename provides path to template file for rules and definitions
                """
    def __init__(self,root_folder, swagger_classifier: SwaggerClassifier, json_generator: JSONGenerator, exported_file_name,app_name,filename = os.path.join(os.path.dirname(__file__),"..","..","Config/Dictionary.xlsx")):
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(filename=root_folder + '/data-classification.log', format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
        self.unclassified_technical_fields = []
        self.all_fields=[]
        self.classification_report={}
        self.parameter_descriptions = swagger_classifier.sh.parameter_descriptions
        self.ce = swagger_classifier.ce
        self.definitions_list = json_generator.definitions_list
        self.workbook = xlsxwriter.Workbook(exported_file_name)
        self.title_formatting = self.workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'purple', 'font_size': 12, 'locked': True})
        self.row_formatting = self.workbook.add_format({'text_wrap': True, 'font_size': 12})
        self.app=app_name
        try:
            self.wb = openpyxl.load_workbook(filename)
            logging.info("Data Dictionary.xlsx file has been loaded")
        except Exception as e:
            logging.error(e)
            sys.exit(e)
        self.rule_list=[]
        

    def generate_dictionary(self,root_folder):
        try:
            self.__generate_details_sheet()
            logging.info("Details Sheet for the Generated Dictionary has been created")
            self.__generate_definitions_sheet()
            logging.info("Definitions Sheet for the Generated Dictionary has been created")
            self.__generate_rules_sheet()
            logging.info("Rules Sheet for the Generated Dictionary has been created")
            self.workbook.close()
            logging.info("Generated Dictionary has been created")
        except Exception as e:
            logging.error(e)
            sys.exit(e)

    def __generate_details_sheet(self):
        worksheet = self.workbook.add_worksheet('Details')
        self.__init_details_sheet_design(worksheet)
        api_directory = []
        definition_list = self.format_api_directory_with_tn()
        definitition_names = self.ce.map_technical_names_to_definintion_names(api_directory, definition_list)
        match_list = []
        row_cusur = 1
        for desc_obj in self.parameter_descriptions:
            api_endpoints = ""
            api_endpoints_array = []
            for api_endpoint_path in self.parameter_descriptions[desc_obj]:
                items= api_endpoint_path["file"].split('/')
                if len(items) > 2:
                    contract = items[1]
                    if api_endpoints == "" and contract not in api_endpoints_array:
                        api_endpoints_array.append(contract)
                        api_endpoints = contract
                    elif contract not in api_endpoints_array:
                        api_endpoints_array.append(contract)
                        api_endpoints = api_endpoints + "; " + contract
            for pd in self.parameter_descriptions[desc_obj]:
                unique_match = desc_obj
                if unique_match not in match_list:
                    try:
                        worksheet.write(row_cusur,0, definitition_names[desc_obj], self.row_formatting)
                    except:
                        self.unclassified_technical_fields.append(desc_obj)
                        print ("Could not find a field type for the parameter {} in row {}".format(desc_obj,row_cusur+1))
                    worksheet.write(row_cusur,1, desc_obj, self.row_formatting)
                    split_obj=desc_obj.split('/',2)
                    if len(split_obj)==2:
                        worksheet.write(row_cusur,2,split_obj[0], self.row_formatting)
                        worksheet.write(row_cusur,3,split_obj[1], self.row_formatting)
                    elif len(split_obj)==1:
                        worksheet.write(row_cusur,3,split_obj[0], self.row_formatting)
                 
                    worksheet.write(row_cusur,4, pd["description"], self.row_formatting)
                    worksheet.write(row_cusur,5, "/"+ self.app+pd["file"], self.row_formatting)
                    if 'endpoint' in pd:
                        worksheet.write(row_cusur,6, pd["endpoint"], self.row_formatting)
                    items= pd["file"].split('/')
                    if len(items)>2:
                        contract=items[1]
                        worksheet.write(row_cusur,7, self.app, self.row_formatting)
                        worksheet.write(row_cusur,8, api_endpoints, self.row_formatting)
                        if contract not in self.classification_report.keys():
                            self.classification_report[contract]=[]
                    isClassified=False
                    if desc_obj not in definitition_names:
                        worksheet.write(row_cusur,9, "No", self.row_formatting)
                        isClassified=False
                    else:
                        worksheet.write(row_cusur,9, "Yes", self.row_formatting)
                        isClassified=True
                    match_list.append(unique_match)
                    row_cusur += 1
                    if contract in self.classification_report.keys():
                        self.classification_report[contract].append({"field":desc_obj,"isClassified": isClassified})
        self.all_fields=match_list

    def format_api_directory_with_tn(self):
        list_definition = []
        list_params = []
        length = len(self.definitions_list)
        for i in range(length):
            exception_values = self.definitions_list[i]["Exception_API_and_Values"]
            for exception_value in exception_values:
                if exception_value != "":
                    list_definition.append(exception_value.lower())
        for parameter in self.parameter_descriptions:
            x = ((self.parameter_descriptions[parameter][0]['file']).rsplit("/", 2)[1])+":"+parameter
            api_name, tn = x.split(":")
            y = api_name[:-1]+":"+tn
            if y.lower() in list_definition:
                list_params.append(x)
            else:
                list_params.append(parameter)
        #list_params = list(set(list_params))
        return list_params

    def __generate_definitions_sheet(self):
        worksheet = self.workbook.add_worksheet('Field Types')
        self.__init_definitions_sheet_design(worksheet)
        details_list = self.ce.dl.definition_list
        row_cusur = 1
        for detail in details_list:
            worksheet.write(row_cusur,0, detail.name, self.row_formatting)
            worksheet.write(row_cusur,1, detail.description, self.row_formatting)
            worksheet.write(row_cusur,2, detail.examples, self.row_formatting)
            worksheet.write(row_cusur,3, detail.sensitive, self.row_formatting)
            worksheet.write(row_cusur,4, detail.guidelines, self.row_formatting)
            row_cusur += 1
     
    def __generate_rules_sheet(self):
        worksheet = self.workbook.add_worksheet('Rules')
        self.rule_list=self.__get_rules()
   
        col_indexes = self.__init_rules_sheet_design(worksheet)
        row_cusur = 1
        for rule in self.rule_list:
            col_cursor = col_indexes['all']
            for item in rule['all']:
                worksheet.write(row_cusur,col_cursor, item, self.row_formatting)
                col_cursor += 1
            col_cursor = col_indexes['any']
            for item in rule['any']:
                worksheet.write(row_cusur,col_cursor, item, self.row_formatting)
                col_cursor += 1
            col_cursor = col_indexes['tags']
            for item in rule['tags']:
                worksheet.write(row_cusur,col_cursor, item, self.row_formatting)
                col_cursor += 1
            row_cusur += 1

    def __init_details_sheet_design(self, worksheet):
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 40)
        worksheet.set_column('C:C', 40)
        worksheet.set_column('D:D', 40)
        worksheet.set_column('E:E', 40)
        worksheet.set_column('F:F', 80)
        worksheet.set_column('G:G', 40)
        worksheet.set_column('H:H', 40)
        worksheet.set_column('I:I', 40)
        worksheet.set_column('J:J', 40)
        worksheet.write('A1', 'Field Type', self.title_formatting)
        worksheet.write('B1', 'Technical Name', self.title_formatting)
        worksheet.write('C1', 'Object', self.title_formatting)
        worksheet.write('D1', 'Field', self.title_formatting)
        worksheet.write('E1', 'Description', self.title_formatting)
        worksheet.write('F1', 'Example File', self.title_formatting)
        worksheet.write('G1', 'Example Endpoint', self.title_formatting)
        worksheet.write('H1', 'Example Product', self.title_formatting)
        worksheet.write('I1', 'Example API', self.title_formatting)
        worksheet.write('J1', 'IsClassified', self.title_formatting)

    def __init_definitions_sheet_design(self, worksheet):
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 80)
        worksheet.set_column('C:C', 40)
        worksheet.set_column('D:D', 20)
        worksheet.set_column('E:E', 80)
        worksheet.write('A1', 'Field Type', self.title_formatting)
        worksheet.write('B1', 'Description', self.title_formatting)
        worksheet.write('C1', 'Examples', self.title_formatting)
        worksheet.write('D1', 'Sensitive', self.title_formatting)
        worksheet.write('E1', 'Guidelines', self.title_formatting)
    #This generates  the rule sheet from the master xlsx  rules 
    def __init_rules_sheet_design(self, worksheet):
        column_cursor = 0
        col_sizes = self.__get_rules_column_sizes()
        #Iterate through the 'all' fields of rules and write it intot the worksheet
        for i in range(col_sizes['all']):
            worksheet.set_column(0,column_cursor,20)
            worksheet.write(0,column_cursor,"All{}".format(i+1), self.title_formatting)
            column_cursor += 1
        #Iterate through the 'any' fields of rules and write it intot the worksheet
        first_any_index = column_cursor
        for i in range(col_sizes['any']):
            worksheet.set_column(0,column_cursor,20)
            worksheet.write(0,column_cursor,"Any{}".format(i+1), self.title_formatting)
            column_cursor += 1
        #Iterate through the 'tags' fields of rules and write it intot the worksheet
        first_tags_index = column_cursor
        for i in range(col_sizes['tags']):
            worksheet.set_column(0,column_cursor,20)
            worksheet.write(0,column_cursor,"Tag{}".format(i+1), self.title_formatting)
            column_cursor += 1
        return {'all': 0, 'any': first_any_index, 'tags': first_tags_index}
        
    def __get_rules_column_sizes(self):
        max_all = 0
        max_any = 0
        max_tags = 0

        for rule in self.rule_list:
            if len(rule['all']) > max_all:
                max_all = len(rule['all'])
            if len(rule['any']) > max_any:
                max_any = len(rule['any'])
            if len(rule['tags']) > max_tags:
                max_tags = len(rule['tags'])
        return {'all': max_all, 'any': max_any, 'tags': max_tags}

    def __get_rules(self):
        rules_list=[]
        rules_sheet =  self.wb['Rules']
        column_indexes = self.__get_rules_column_indexes()
        for i in range(2, rules_sheet.max_row+1):
            rule = {'any': [], 'all': [], 'tags': [], 'scope':[]}  
            for c in column_indexes['any']:
                if rules_sheet.cell(row = i, column = c).value != None:
                    rule['any'].append(rules_sheet.cell(row = i, column = c).value)
            for c in column_indexes['all']:
                if rules_sheet.cell(row = i, column = c).value != None:
                    rule['all'].append(rules_sheet.cell(row = i, column = c).value)
            for c in column_indexes['tags']:
                if rules_sheet.cell(row=i,column=c).value!=None:
                    rule['tags'].append(rules_sheet.cell(row = i, column = c).value)
            rules_list.append(rule)
          
        return rules_list

    def __get_rules_column_indexes(self):
        index_dict = {'any': [], 'all': [], 'tags': []}
        rules_sheet =  self.wb['Rules']
        for i in range(1, rules_sheet.max_column+1):
            column_name = rules_sheet.cell(row = 1, column = i).value
            if 'Any' in column_name:
                index_dict['any'].append(i)
            elif 'All' in column_name:
                index_dict['all'].append(i)
            elif 'Tag' in column_name:
                index_dict['tags'].append(i)
        return index_dict

